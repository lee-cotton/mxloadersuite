from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from mxls.MXHttpRequest import ConnectionDetails
from mxls.MXHttpRequest import MXHttpRequest
from mxls.MXLoaderWorksheet import MXLoaderWorkSheet
from mxls.logging.MXLogger import MXLogger
 
class Configuration(object):
     def __init__(self, configWorkSheet):
        self.baseLanguage = configWorkSheet['B16'].value
        self.translationLanguage = configWorkSheet['B17'].value
        self.nameSpace = configWorkSheet['B35'].value
        self.startFromRow = configWorkSheet['B31']
          
class MXLoader(object):

    def __init__(self, connectionDetails):
        self.connectionDetails = connectionDetails
        self.workSheets = []
        self.loadWorkBook
        self.logger = MXLogger.getLogger()

    def loadWorkSheet(self, workBook, sheetName):

        workSheet = workBook[sheetName]

        objectStructure = workSheet['A1'].value
        objectName = workSheet['B1'].value
        action = workSheet['C1'].value
     
        self.logger.info('Object Structure=' + objectStructure)
        self.logger.info('Object Name=' + objectName)
        self.logger.info('Action=' + action)
        self.logger.info('Start From Line=' )

        mxLoaderWorkSheet = MXLoaderWorkSheet()
        mxLoaderWorkSheet.setObjectStructure(objectStructure)
        mxLoaderWorkSheet.setObjectName(objectName)
        mxLoaderWorkSheet.setAction(action)
        mxLoaderWorkSheet.setBaseLanguage(self.configuration.baseLanguage)
        mxLoaderWorkSheet.setTranslationLanguage(self.configuration.translationLanguage)
        mxLoaderWorkSheet.setNameSpace(self.configuration.nameSpace)

        mxLoaderWorkSheet.setWorkSheet(workSheet)

        mxLoaderWorkSheet.parseWorkSheet()

        request = mxLoaderWorkSheet.toXml()

        if request == None:
            self.logger.info('No rows loaded from worksheet:' + sheetName)
            return

        httpRequest = MXHttpRequest(self.connectionDetails)

        httpRequest.sendRequest(objectStructure, request) 

    def loadWorkBook(self, excelFile):

        #if the file name starts with '~$' then this is a temporay Excel file so it's skipped.
        if excelFile.name.startswith('~$'):
            return

        self.logger.info('Loading Workbook: ' + excelFile.name)

        wb = load_workbook(excelFile)

        configWorkSheet = wb['Config']

        self.configuration = Configuration(configWorkSheet)

        for sheetName in wb.sheetnames:

            #You can create a Worksheet called 'Stop' to prevent subsequent worksheets being loaded from the current file.
            if sheetName == 'Stop':
                self.logger.info('Encountered worksheet: Stop. Processing of current file is terminated.')
                break

            workSheet = wb[sheetName]

            isHidden = workSheet.sheet_state == 'veryHidden'

            if isHidden or sheetName == "About" or sheetName == "Config":
                continue
 
            self.logger.info('Loading Worksheet: ' + sheetName)

            self.loadWorkSheet(wb, sheetName)