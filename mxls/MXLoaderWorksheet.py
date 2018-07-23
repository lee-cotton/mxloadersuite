from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from mxls.MXNode import MXNode
from mxls.MXAttribute import MXAttribute
from mxls.logging.MXLogger import MXLogger

class MXLoaderWorkSheet(object):

    def __init__(self):
        self.logger = MXLogger.getLogger()

        self.rootNode = None
        self.parentNode = None
        self.currentNode = None
        self.attributes = []
        self.previousObjectFullName = ''

    def setObjectStructure(self, objectStructure):    
        self.objectStructure = objectStructure
    def getObjectStructure(self):
        return self.objectStructure
    def setObjectName(self, objectName):    
        self.objectName = objectName
    def getObjectName(self):
        return self.objectName

    def setAction(self, action):    
        self.action = action.split('-')[1]
        self.operation = action.split('-')[0]
    def getAction(self):
        return self.action

    def setBaseLanguage(self, baseLanguage):    
        self.baseLanguage = baseLanguage
    def getBaseLanguage(self):
        return self.baseLanguage

    def setTranslationLanguage(self, translationLanguage):    
        self.translationLanguage = translationLanguage
    def getTranslationLanguage(self):
        return self.translationLanguage

    def setNameSpace(self, nameSpace):    
        self.nameSpace = nameSpace
    def getNameSpace(self):
        return self.nameSpace

    def setWorkSheet(self, workSheet):
        self.workSheet = workSheet

    def loadAttributes(self):
        self.logger.debug('Attributes: ')

        for row in self.workSheet.iter_cols(min_row=2, max_row=2):
            for cell in row:
                if cell.value == None:
                    break
                
                cellValue = cell.value
                cellValue = cellValue.replace('\n', "") 

                self.attributes.append(cellValue)
                
                self.logger.debug('Attribute Name: ' + cellValue)

    def xhAdd(self, attributeName, attributeValue):

        if attributeValue == '' or attributeValue == None:
            return
        
        position = attributeName.rfind('.')

        currentObjectFullName = attributeName[:position]

        currentAttributeName = attributeName[position+1:]

        position = currentObjectFullName.rfind('.')

        if position < 0:
            currentObjectName = currentObjectFullName
        else:
            currentObjectName = currentObjectFullName[position+1:]
        
        if currentObjectFullName != self.previousObjectFullName: 
            self.parentNode = self.parentNode.addChildNodeNE(self.currentNode)
            self.currentNode = MXNode(currentObjectName, self.parentNode)

        currentValue = MXAttribute(currentAttributeName, attributeValue)

        self.currentNode.addAttribute(currentValue) 

        self.previousObjectFullName = currentObjectFullName

    def parseWorkSheet(self):

        self.loadAttributes()

        self.rootNode = MXNode(self.objectName + 'Set', None)
    
        for row in self.workSheet.iter_rows(min_row=3):

            #Stop when an empty row is reached.
            if row[0].value is None and row[1].value is None:
                break

            self.parentNode = self.rootNode
            self.currentNode = MXNode(self.objectName, self.parentNode)

            self.previousObjectFullName = self.objectName

            self.currentNode.setSheetRow(row)

            i=0

            for attribute in self.attributes:
                self.xhAdd(self.objectName + '.' + attribute, row[i].value)
                i = i+1    
            
            self.parentNode.addChildNodeNE(self.currentNode)
               

    def toXml(self):
        
        #No rows loaded so return None
        if not self.rootNode.children:  
            return None

        operation = self.operation
        objectStructure = self.objectStructure
        nameSpace = self.nameSpace
        baseLanguage = self.getBaseLanguage()
        translationLanguage = self.getTranslationLanguage()

        attr = ""

        xml = '<' + operation + objectStructure
        xml = xml + ' xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"'
        xml = xml + ' xmlns="' + nameSpace + '"'
        xml = xml + ' baseLanguage="' + baseLanguage + '"'
        xml = xml + ' transLanguage="' + translationLanguage + '">'
        xml = xml + attr

        xml = xml + '<' + self.objectStructure + 'Set action="' + self.action + '">'

        xml = xml + self.rootNode.toXml()

        xml = xml + '</' + self.objectStructure + 'Set>'

        xml = xml + '</' + operation + objectStructure + '>'
   
        return xml
