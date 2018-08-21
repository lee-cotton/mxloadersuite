from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time

def writeRow(ws1, row, rowNumber):

    for index, column in enumerate(row):
        cellValue = column
        cellValue = cellValue.strip('\n')
        cellValue = cellValue.replace('\n', ' ')
        cellValue = cellValue.strip('"')
        cellValue = cellValue.encode('unicode_escape').decode('utf-8')
        
        if rowNumber > 1 and (index==2 or index == 6 or index == 7):
            cellValue = time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime(time.mktime(time.strptime(cellValue, "%m/%d/%Y %H:%M:%S"))))

        ws1.cell(column=(index+1), row=rowNumber, value=cellValue)
        ws1.cell(column=(index+1), row=rowNumber).alignment = Alignment(wrapText=True)
    #if rowNumber > 100:
    #    raise ValueError('Max rows')

wb = Workbook()
dest_filename = 'Compass Defects.xlsx'
ws1 = wb.active
ws1.title = "Defects"
rowNumber = 1
lineNumber = 1

#try:

with open('COMPASS_PARTIAL.csv') as f:
#with open('MAXIMO_TEMP_EXPORT_IBM.csv') as f:
    row = None
    for line in f:

        print(lineNumber)
        lineNumber = lineNumber + 1

        if row != None and len(row) == 8:
            row = None
                
        if line.count('|') < 7:   # Note continuation
            if row == None:
                row = line.split('|')
                continue
            if line.count('|') == 0:
                row[-1] += line
                continue
            else:
                temp = line.split('|')

                x=0
                for g in temp:
                    if x == 0:
                        row[-1] += g.strip('\n')
                    else:
                        row.append(g.strip('\n'))
                    x += 1
                if len(row) == 8:
                    writeRow(ws1, row, rowNumber)
                    rowNumber += 1
                    continue

                continue

        row = line.split('|')

        if len(row) == 8:
            writeRow(ws1, row, rowNumber)
            rowNumber += 1
            continue

#except ValueError:
#   print('Reached max rows')

wb.save(filename = dest_filename)